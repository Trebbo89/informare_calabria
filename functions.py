from tkinter import filedialog as fd
from tkinter.messagebox import showinfo
from constants import *
import pandas as pd
import os
from datetime import datetime
import customtkinter as ctk
import inspect
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter

# il metodo serve per generare una cartella contenente i log di eventuali errori
# app -> istanza CustomTKinter
# error_message -> messaggi di errore
def generate_log_errore(app, error_message):
    
    try:
        current_path = os.getcwd()
        path_log_dir = current_path + '/log'
        
        if not os.path.exists(path_log_dir):
            os.mkdir(path_log_dir)
            
        stack = inspect.stack()
        metodo_chiamante = stack[1].function
        
        current_date_and_hour = f"[{datetime.now().strftime(PATTERN_DD_MM_YY_WITH_TIME_AND_SECONDS)}]"
        complete_error_message = f'{current_date_and_hour} - [ERRORE] {metodo_chiamante}() - {error_message}\n'
            
        with open(f'{path_log_dir}/log_errori.txt', 'a') as f:
            f.writelines(str(complete_error_message))
    except Exception as e:
        print(e)

# la funzione serve per generare righe o colonne del template grid
# app -> istanza customTKinter
# number_of_row_or_column -> numero di colonne o righe da generare
# generate_row -> indica se generare righe (True) o colonne (False)
def configuration_row_or_column_for_grid_template(app, number_of_row_or_column = 0, generate_row = False):
    try:
        for i in range(number_of_row_or_column):
            if generate_row:
                app.grid_rowconfigure(i, weight=1)
            else:
                app.grid_columnconfigure(i, weight=1)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
            
# operazioni da eseguire quando si preme sulla checkbox
# app -> istanza CustomTKinter
# option_menu -> option contenente gli id degli studenti
# button_insert_update_studente -> bottone di inserimento/aggiornamento studente
# checkbox_value -> valore della checkbox (on=selezionata) (off=deselezionata)
# course_dict_with_students -> dizionario contenente il nome del corso e gli studenti
def operations_flag_checkbox(app, option_menu, button_insert_update_studente, checkbox_value, course_dict_with_students, button_delete):
    try:
        enable_or_disabled_option_id_studente(app, option_menu, checkbox_value)
        chage_text_button_insert_update_student(app, button_insert_update_studente, checkbox_value)
        set_values_options_id_students(app, option_menu, course_dict_with_students)
        enable_or_disable_delete_button(app, button_delete, checkbox_value, option_menu.get())
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo si occupa di abilitare e disabilitare l'option per la selezione dell'id dello studente
# app -> istanza CustomTKinter
# option_menu -> option contenente gli id degli studenti
# checkbox_value -> valore della checkbox (on=selezionata) (off=deselezionata)
def enable_or_disabled_option_id_studente(app, option_menu, checkbox_value):
    try:
        if checkbox_value == 'off':
            option_menu.configure(state='disabled')
        else:
            option_menu.configure(state='')
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
# il metodo si occupa di aggiornare il testo del pulsante di inserimento/aggiornamento degli studenti
# app -> istanza CustomTKinter
# button_insert_update_studente -> bottone di inserimento/aggiornamento studente
# checkbox_value -> valore checkbox
def chage_text_button_insert_update_student(app, button_insert_update_studente, checkbox_value):
    try:
        button_insert_update_studente.configure(text=TEXT_INSERT_STUDENT if checkbox_value == 'off' else TEXT_UPDATE_STUDENT)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo si occupa di generare la dropdown con la lista di tutti gli id degli studenti
# app -> istanza CustomTKinter
# option_menu -> oggetto option
# course_dict_with_students -> dizionario contenente il nome del corso e gli studenti
def set_values_options_id_students(app, option_menu, course_dict_with_students):
    
    try:
        if 'students' in course_dict_with_students.keys() and course_dict_with_students['students']:
            students_id_list = [str(student['id']) for student in course_dict_with_students['students']]
            
            option_menu.configure(values=[OPTION_ID_STUDENTE_DEFAULT] + students_id_list)
            
            if len(students_id_list) <= 0:
                option_menu.set(OPTION_ID_STUDENTE_DEFAULT)
            elif option_menu.get() != OPTION_ID_STUDENTE_DEFAULT and int(option_menu.get()) > int(students_id_list[-1]):
                option_menu.set(students_id_list[-1])
        else:
            option_menu.configure(values=[OPTION_ID_STUDENTE_DEFAULT])
            option_menu.set(OPTION_ID_STUDENTE_DEFAULT)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo, se richiamato da un bottone, serve per trasformarlo in un file input
# app -> istanza CustomTKinter
# filetypes -> tipologia di file da poter selezionare nell'input
def file_selected(app, file_types = FILETYPES_TXT):
    
    try:
        filetypes = (
            file_types,
        )
        
        filename = fd.askopenfilename(
            title='Apri un file',
            initialdir='./',
            filetypes=filetypes
        )
        
        if filename:
            showinfo(title='Seleziona un file', message=filename)
            return filename
        
        return ''
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo presenta tutte le operazioni da svolgere una volta cliccato il pulsante di inserimento/aggiornamento studente
# app -> istanza customTKinter
# ctk -> oggetto customTkinter contenente tutti i vari input
# course_dict_with_students -> dizionario contenente il nome del corso e gli studenti
# course_name -> nome del corso
# student -> studente da inserire nel dizionario
# textarea -> è l'oggetto textarea che ci servirà successivamente per inserire i valori all'interno della stessa
# checkbox_value -> valore checkbox
# id_student_selected -> id studente selezionato
def operations_insert_update_student(app, ctk, course_dict_with_students = dict(), course_name = '', student = '', textarea = None, checkbox_value = '', option_menu = None):
    
    try:
        result = check_field(app, course_name, student)
        
        if result:
            set_value_dict_course_and_students(app, course_dict_with_students, course_name, student, checkbox_value, option_menu)
            set_value_dict_in_textbox(app, textarea, course_dict_with_students, course_name)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo verica i campi nome corso e studente
# app -> istanza customTKinter
# ctk -> oggetto customTkinter contenente tutti i vari input
# course_name -> Nome corso
# student -> nome e cognome studente   
def check_field(app, course_name = '', student = ''):
    
    try:
        if course_name == '' or student == '':
            message = f"Inserire il valore per il campo \n{'Nome Corso' if course_name == '' else 'Nome Cognome studente'}"
            generate_modal(app, message)
            return False
        
        return True
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo genera una modal
# app -> istanza customTKinter
# ctk -> oggetto customTkinter contenente tutti i vari input
# message -> messaggio da visualizzare nella modal
def generate_modal(app, message, title = 'Errore'):
    
    try:
        # Creazione di una finestra di dialogo
        dialog = ctk.CTkToplevel(app)
        dialog.geometry('300x150')
        dialog.title(title)
        
        configuration_row_or_column_for_grid_template(dialog, 2, True)
        configuration_row_or_column_for_grid_template(dialog, 1, False)
        
        # Imposta la finestra di dialogo come modale
        dialog.transient(app)  # Imposta la finestra come figlia della finestra principale
        dialog.grab_set()  # Blocca le interazioni con la finestra principale
        
        # Messaggio della dialog
        label = ctk.CTkLabel(dialog, text=message)
        label.grid(pady=PADDING_Y, padx=PADDING_X, sticky='we')
        
        # Pulsante per chiudere la dialog
        close_button = ctk.CTkButton(dialog, text="OK", command= dialog.destroy)
        close_button.grid(pady=20, padx=20, sticky='we')
        
        # Assicurati che la finestra di dialogo sia sopra la principale
        dialog.wait_window(dialog)  # Aspetta che la finestra venga chiusa
    except Exception as e:
        generate_log_errore(app, e)
    
# il metodo inserisce i valori all'interno del dizionario per farli successivamente visualizzare nella textbox
# app -> istanza CustomTKinter
# course_dict_with_students -> dizionario contenente il nome del corso e gli studenti
# course_name -> nome del corso
# student -> studente da inserire nel dizionario
# checkbox_value -> valore checkbox
# option_menu -> oggetto option contenente gli id degli studenti
def set_value_dict_course_and_students(app, course_dict_with_students = dict(), course_name = '', student = '', checkbox_value = '', option_menu = None):
    
    try:
        if 'course_name' in course_dict_with_students.keys() and course_name != course_dict_with_students['course_name']:
                course_dict_with_students['course_name'] = course_name
                
        if checkbox_value == 'off':
            
            if ('course_name' in course_dict_with_students.keys() and course_dict_with_students['course_name']) and ('students' in course_dict_with_students.keys() and course_dict_with_students['students']):
                last_id = course_dict_with_students['students'][-1]['id']
                course_dict_with_students['students'].append({'id': int(last_id + 1), 'student': student})
            else:
                course_dict_with_students['course_name'] = course_name
                course_dict_with_students['students'] = [{'id': 1, 'student': student}]
                
            set_values_options_id_students(app, option_menu, course_dict_with_students)
                
            
        elif checkbox_value == 'on' and (option_menu != None and option_menu.get() != '' and option_menu.get() != OPTION_ID_STUDENTE_DEFAULT):
            if 'course_name' in course_dict_with_students.keys() and course_dict_with_students['students']:
                for student_obj in course_dict_with_students['students']:
                    
                    if student_obj['id'] == int(option_menu.get()):
                        student_obj['student'] = student
                
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# inserisce il valore del dizionario all'interno della textbox
# app -> istanza CustomTKinter
# textarea -> oggetto textarea dove inserire il corso e gli studenti
# course_dict_with_students -> dizionario con corso e studenti
# course_name -> nome del corso
def set_value_dict_in_textbox(app, textarea, course_dict_with_students, course_name):
    
    try:
        
        if 'course_name' in course_dict_with_students.keys() and course_dict_with_students['course_name']:
            
            textarea.configure(state='normal')
            textarea.delete('1.0', 'end')
            textarea.insert('1.0', course_name)
            textarea.insert('2.0', '\n')
            
            for index, value in enumerate(course_dict_with_students['students']):
                textarea.delete(f'{int(index + 3)}.0', 'end')
                textarea.insert(f'{int(index + 3)}.0', f'{value["id"]} - {value["student"]}\n')
                
            textarea.configure(state='disabled')
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
    
    
# il metodo si occupa di abilitare e disabilitare il pulsante di eliminazione di uno studente
# app -> istanza CustomTKinter
# button_delete -> oggetto CTkButton
# checkbox_value -> valore della checkbox
# option_value -> valore dell'id studente selezionato
def enable_or_disable_delete_button(app, button_delete, checkbox_value, option_value):
    try:
        if checkbox_value == 'on' and (option_value != None and option_value != '' and option_value != OPTION_ID_STUDENTE_DEFAULT):
            button_delete.configure(state='normal')
        else:
            button_delete.configure(state='disabled')
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo si occupa di eliminare lo studente in base all'id selezionato nell'option menu
# app -> istanza CustomTKinter
# textarea -> oggetto textarea dove inserire il corso e gli studenti
# course_dict_with_students -> dizionario con corso e studenti
# option_menu -> oggetto option contenente gli id degli studenti
def delete_student(app, textarea, course_dict_with_students, option_menu):
    
    try:
        if 'course_name' in course_dict_with_students.keys() and course_dict_with_students['course_name']:
            course_name = course_dict_with_students['course_name']
        
        if 'students' in course_dict_with_students.keys() and course_dict_with_students['students'] and option_menu.get() != OPTION_ID_STUDENTE_DEFAULT:
            new_students_list = list()
            count_id_student = 1
            
            for student in course_dict_with_students['students']:
                if int(student['id']) != int(option_menu.get()):
                    new_students_list.append({'id': count_id_student, 'student': student['student']})
                    count_id_student += 1
                    
            course_dict_with_students['students'] = new_students_list
            set_value_dict_in_textbox(app, textarea, course_dict_with_students, course_name)
            set_values_options_id_students(app, option_menu, course_dict_with_students)
        elif 'students' in course_dict_with_students.keys() and len(course_dict_with_students['students']) <= 0:
            option_menu.set(OPTION_ID_STUDENTE_DEFAULT)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo si occupa di gestire tutte le operazioni quando si clicca sul pulsante di importazione studenti
# app -> istanza CustomTKinter
# textarea -> oggetto textarea 
# course_dict_with_students -> dizionario con corso e studenti
def operations_import_students_from_file(app, textarea, course_dict_with_students, course_input):
    
    try:
        filename = file_selected(app, FILETYPES_XLSX)
        
        if filename != '':
            import_students_from_file(app, filename, textarea, course_dict_with_students, course_input)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per importare gli studenti da un foglio excel al programma
# app -> istanza CustomTKinter
# filename -> percorso con il nome del file
# textarea -> oggetto textarea
def import_students_from_file(app, filename, textarea, course_dict_with_students, course_input = None):
    
    try:
        if filename:
            df = pd.read_excel(filename, sheet_name=None, engine='openpyxl')
            df_sheet = df[list(df.keys())[0]]
            course_name = df_sheet['NOME CORSO'][0]
            students_list = df_sheet['STUDENTI'].values
            
            course_dict_with_students['course_name'] = course_name
            course_dict_with_students['students'] = [ {'id': i+1, 'student': student} for i, student in enumerate(students_list) ]
            
            if textarea != None:
                set_value_dict_in_textbox(app, textarea, course_dict_with_students, course_name)
            
            if course_input != None:
                course_input.delete(0, 'end')
                course_input.insert(0, course_name)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per generare il file del corso
# app -> istanza CustomTKinter
# course_dict_with_students -> dizionario con corso e studenti
def generate_course_file(app, course_dict_with_students):
    
    try:
        
        current_path = os.getcwd()
        path_generation_xlsx_course_file = current_path + '/file generazione corso'
        
        if not os.path.exists(path_generation_xlsx_course_file):
            os.mkdir(path_generation_xlsx_course_file)
            
        course_name = course_dict_with_students["course_name"]
        excel_filename_and_path = f'{path_generation_xlsx_course_file}/studenti_corso_{course_name}.xlsx'
        
        course_dict_with_students_excel = dict()
        course_dict_with_students_excel['NOME CORSO'] = [course_name if i == 0 else '' for i in range(0, len(course_dict_with_students["students"]))]
        course_dict_with_students_excel['STUDENTI'] = [student['student'] for student in course_dict_with_students['students']]
        
        df = pd.DataFrame(course_dict_with_students_excel)
        df.to_excel(excel_filename_and_path, index=False, sheet_name=course_name)
        
        wb = load_workbook(excel_filename_and_path)
        ws = wb.active
        
        ws.column_dimensions['A'].width = len(course_name) + 10
        ws.column_dimensions['B'].width = max([len(student) for student in course_dict_with_students_excel['STUDENTI']]) + 10
        
        fill_blue = PatternFill(start_color='9BC2E6', end_color='9BC2E6', fill_type='solid')
        
        ws['A1'].fill = fill_blue
        ws['B1'].fill = fill_blue
        
        border = Border(
            left=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            right=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            top=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            bottom=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
        )
        
        set_border_column_excel(app, ws, ['A', 'B'], border)
        
        wb.save(excel_filename_and_path)
        generate_modal(app, 'File generato con successo', 'Successo')
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
    
# il metodo serve per settere alle varie colonne dell'excel il bordo
# app -> istanza CustomTKinter
# ws -> worksheet
# columns_list -> lista delle colonne a cui settare il bordo
# border -> tipologia bordo da settare
def set_border_column_excel(app, ws, columns_list = [], border = Border()):
    
    try:
        
        for column in columns_list:
            for cell in ws[column]:
                
                if cell.value in LIST_TEXT_WITHOUT_BORDER:
                    if cell.value == TEXT_NO_BORDER:
                        cell.value = ''
                    continue
                
                cell.border = border
            
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        raise
        
        
# il metodo gestisce le operazioni per l'importazione del file di report di mynaparrot
# app -> istanza CustomTKinter
def operations_import_report_mynaparrot(app, imported_students_course):
    try:
        filename = file_selected(app, FILETYPES_XLSX)
        
        if filename:
            read_report_mynaparrot(app, filename, imported_students_course)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo legge il report mynaparrot ed estrai i dati necessari
# app -> istanza CustomTKinter
# filename -> percorso del file
def read_report_mynaparrot(app, filename, imported_students_course):
    
    try:
        dict_for_export_custom_excel_report = dict()
        
        df = pd.read_excel(filename, sheet_name=None, engine='openpyxl')
        
        sheet_room = list(df.keys())[0]
        sheet_user = list(df.keys())[1]
        
        df_sheet_room = df[sheet_room]
        
        title_label =  list(df_sheet_room.keys())[0]
        course_name = list(df_sheet_room.keys())[1]
        
        date_created_course = ''
        date_ended_course = ''
        
        for index, element in enumerate(df_sheet_room[title_label]):
            if element == 'Created':
                date_created_course = datetime.fromisoformat(df_sheet_room.loc[df_sheet_room.index[index], course_name]).strftime(PATTERN_DD_MM_YY_WITH_TIME)
            elif element == 'Ended':
                date_ended_course = datetime.fromisoformat(df_sheet_room.loc[df_sheet_room.index[index], course_name]).strftime(PATTERN_DD_MM_YY_WITH_TIME)
                
        date_created_course = date_and_time_adjustment(app, date_created_course, 15)
        date_ended_course = date_and_time_adjustment(app, date_ended_course, 15)
        
        minutes_course_attendance = calculation_of_course_minutes(app, date_created_course, date_ended_course)
        
        dict_for_export_custom_excel_report['created_course_date'] = date_created_course
        dict_for_export_custom_excel_report['ended_course_date'] = date_ended_course
        dict_for_export_custom_excel_report['minutes_course_attendence'] = minutes_course_attendance
        
        df_sheet_user = df[sheet_user][['Name', 'Moderator', 'Joined', 'Left']]
        
        moderators = list(df_sheet_user[df_sheet_user['Moderator'] == 'Yes']['Name'].values)
        print(moderators)
        
        dict_for_export_custom_excel_report['moderators'] = moderators
        
        students_without_moderators = df_sheet_user[df_sheet_user['Moderator'] == 'No']
        
        students_list = generate_students_array_for_excel(app, students_without_moderators)
        
        students_date_time_adjustment(app, students_list, date_created_course, date_ended_course)
        
        for student in students_list:
            minutes = calculation_of_course_minutes(app, student['access'], student['left'])
            student['minutes'] = minutes
            
        students_list_for_excel = {
            'Nome e Cognome': [student['name'] for student in students_list],
            'Entrata': [student['access'] for student in students_list],
            'Uscita': [student['left'] for student in students_list],
            'Minuti': [student['minutes'] for student in students_list]
        }
        
        # lascio una riga vuota per poi inserire gli utenti assenti
        add_blank_row_for_excel(app, students_list_for_excel, 1)
        
        # controllo utenti assenti
        check_absent_students(app, students_list_for_excel, imported_students_course, minutes_course_attendance)
        
        # lascio 7 righe vuote dopo l'inserimento degli studenti assenti
        # add_blank_row_for_excel(app, students_list_for_excel, 7)
        
        generate_row_without_border(app, students_list_for_excel, 1)
        
        # inserimento sezione materia
        time = get_time_from_date_created_and_ended_course(app, date_created_course, date_ended_course)
        insert_details_course_section(app, students_list_for_excel, course_name, time, moderators)
        
        generate_row_without_border(app, students_list_for_excel, 1)
        
        insert_note_section(app, students_list_for_excel)
        generate_row_without_border(app, students_list_for_excel, 1)
        
        insert_totals_and_sign(app, students_list_for_excel)
        
        generate_excel_report(app, students_list_for_excel, course_name, time, moderators, filename)
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per adeguare l'ora in base alla tolleranza inserita
# app -> istanza CustomTKinter
# date -> data e ora da modificare
# tollerability -> tollerabilità
def date_and_time_adjustment(app, date, tollerability):
    
    try:
        date_obj = datetime.strptime(date, PATTERN_DD_MM_YY_WITH_TIME)
        
        date_event = date_obj.strftime(PATTERN_DD_MM_YY)
        
        hour = date_obj.strftime('%H')
        minutes = date_obj.strftime('%M')
        
        if int(minutes) > int(tollerability):
            hour = int(hour) + 1
            
            if hour >= 0 and hour <= 9:
                hour = '0' + str(hour)
            else:
                hour = str(hour)
        
        minutes = '00'
        
        new_date_and_hour = f'{date_event} {hour}:{minutes}'
        
        return new_date_and_hour
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo si occupa di calcolare i minuti di frequentazione del corso
# app -> istanza CustomTKinter
# date1 -> data minore (10-10-2024 15:00)
# date2 -> data maggiore (10-10-2024 18:00)
def calculation_of_course_minutes(app, date1, date2):
    
    try:
        start_date = datetime.strptime(date1, PATTERN_DD_MM_YY_WITH_TIME)
        end_date = datetime.strptime(date2, PATTERN_DD_MM_YY_WITH_TIME)
        
        difference = end_date - start_date
        
        difference_in_minutes = int(difference.total_seconds() / 60)
        
        return difference_in_minutes
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per creare un array contenente tutti gli studenti del corso controllando disconnessioni e assenze
# app -> istanza CustomTKinter
# students_without_moderators -> dataframe di studenti senza moderatore
def generate_students_array_for_excel(app, students_without_moderators):
    
    try:
        
        students_list = list()
        
        for i in range(0, len(students_without_moderators)):
            student = students_without_moderators.iloc[i]
            
            all_student_access = student_access_control(app, student['Joined'], student['Left'])
            
            for date in all_student_access:
                student_details = {
                    'name': student['Name'],
                    'access': date['access'],
                    'left': date['left']
                }
                students_list.append(student_details)
                
    
        return students_list
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per verificare gli accessi dello studente al corso
# app -> istanza CustomTKinter
# student_access_dates -> date di accesso dello studente
# student_left_dates -> date di uscita dello studente
def student_access_control(app, student_access_dates = '', student_left_dates = '' ):
    
    try:
        all_student_access = list()
        student_access = dict()
        
        student_access_dates_split = [ date for date in student_access_dates.split('\n') if date ]
        student_left_dates_split = [ date for date in student_left_dates.split('\n') if date ]
        
        absent_user_beyond_tolerability = False
        
        for i in range(0, len(student_access_dates_split)):
            
            if student_access_dates_split[i] != None and student_access_dates_split[i] != '':
                student_access_date = datetime.fromisoformat(student_access_dates_split[i]).strftime(PATTERN_DD_MM_YY_WITH_TIME)
                student_left_date = datetime.fromisoformat(student_left_dates_split[i]).strftime(PATTERN_DD_MM_YY_WITH_TIME)
                
                if i == 0 or absent_user_beyond_tolerability:
                    student_access['access'] = student_access_date
                    absent_user_beyond_tolerability = False
                    
                if i < ( len(student_access_dates_split) - 1 ):
                    next_access = datetime.fromisoformat(student_access_dates_split[i + 1]).strftime(PATTERN_DD_MM_YY_WITH_TIME)
                    difference_minutes_access_left_course = calculation_of_course_minutes(app, student_left_date, next_access)
                    
                    if difference_minutes_access_left_course >= MINUTI_TOLLERANZA_STUDENTE:
                        student_access['left'] = student_left_date
                        all_student_access.append(student_access)
                        student_access = dict()
                        absent_user_beyond_tolerability = True
                        
                elif i >= ( len(student_access_dates_split) - 1 ):
                    student_access['left'] = student_left_date
                    all_student_access.append(student_access)
                    student_access = dict()
                
                    
        return all_student_access
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo controlla le date e l'ora e le aggiusta in base alla data di creazione e fine corso
# app -> istanza CustomTKinter
# students_list -> lista degli studenti
# creation_date_course -> data di creazione del corsoù
# ending_date_course -> data fine corso
def students_date_time_adjustment(app, students_list, creation_date_course, ending_date_course):
    
    try:
        for student in students_list:
            greather_date_access, minor_date_access = check_date_is_greather(app, student['access'], creation_date_course)
            greather_date_left, minor_date_left = check_date_is_greather(app, student['left'], ending_date_course)
            
            access_minutes_difference = calculation_of_course_minutes(app, minor_date_access, greather_date_access)
            left_minutes_difference = calculation_of_course_minutes(app, minor_date_left, greather_date_left)
            
            if access_minutes_difference <= MINUTI_TOLLERANZA_ACCESSO_AL_CORSO:
                student['access'] = creation_date_course
            
            if left_minutes_difference <= MINUTI_TOLLERANZA_USCITA_DAL_CORSO:
                student['left'] = ending_date_course
                
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo controlla quale data è maggiore e le restituisce
# app -> istanza CustomTKinter
# date1_str -> data1 in formato string (10-10-2024 14:59)
# date2_str -> data2 in formato string (10-10-2024 14:59)
def check_date_is_greather(app, date1_str, date2_str):
    
    try:
        date1 = datetime.strptime(date1_str, PATTERN_DD_MM_YY_WITH_TIME)
        date2 = datetime.strptime(date2_str, PATTERN_DD_MM_YY_WITH_TIME)
        
        if date1 > date2:
            return date1.strftime(PATTERN_DD_MM_YY_WITH_TIME), date2.strftime(PATTERN_DD_MM_YY_WITH_TIME)
        elif date1 < date2:
            return date2.strftime(PATTERN_DD_MM_YY_WITH_TIME), date1.strftime(PATTERN_DD_MM_YY_WITH_TIME)
        else:
            return date1.strftime(PATTERN_DD_MM_YY_WITH_TIME), date2.strftime(PATTERN_DD_MM_YY_WITH_TIME)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo si occupa di gestire tutte le operazioni quando si clicca sul pulsante di importazione studenti
# app -> istanza CustomTKinter
# input_text_filaname -> oggetto input 
# button_import_report -> oggetto bottone per importare il report mynaparrot
# all_students_course -> lista di studenti da popolare
def operations_input_file_get_students(app, input_text_filaname, button_import_report, all_students_course):
    
    try:
        filename = file_selected(app, FILETYPES_XLSX)
        
        if filename:
            input_text_filaname.configure(state='normal')
            input_text_filaname.delete(0, 'end')
            input_text_filaname.insert(0, filename[(filename.rfind('/') + 1):])
            input_text_filaname.configure(state='disabled')
            button_import_report.configure(state='normal')
            course_details = dict()
            import_students_from_file(app, filename, None, course_details, None)
            for student in course_details['students']:
                all_students_course.append(student['student'])
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo controlla gli studenti assenti
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale
# imported_students_course -> studenti recuperati dal file inserito in precedenza
# course_duration -> durata del corso
def check_absent_students(app, students_list_for_excel, imported_students_course, course_duration):
    
    try:
        
        for imported_student in imported_students_course:
            absent_student = True
            for course_student in students_list_for_excel['Nome e Cognome']:
                if imported_student == course_student:
                    absent_student = False
                    break
                
            if absent_student:
                insert_row_excel_list(app, students_list_for_excel, imported_student, TEXT_UTENTE_ASSENTE, TEXT_UTENTE_ASSENTE, course_duration)
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per aggiungere le righe vuote all'excel
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale
# number_black_row -> numero di righe vuote da inserire
def add_blank_row_for_excel(app, students_list_for_excel, number_black_row):
    
    try:
        
        for i in range(0, number_black_row):
            insert_row_excel_list(app, students_list_for_excel, '', '', '', '')
            
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo inserisce la sezione relativa al dettaglio del corso
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale
# course_name -> nome del corso
# date_created_course -> data inizio corso
# date_ended_course -> data fine corso
# moderators -> moderatori
def insert_details_course_section(app, students_list_for_excel, course_name, time, moderators):
    
    try:
        
        insert_row_excel_list(app, students_list_for_excel, TEXT_MATERIA, '', TEXT_ORE, TEXT_DOCENTE)
        # insert_row_excel_list(app, students_list_for_excel, course_name, '', time, moderators_str)
        for i, moderator in enumerate(moderators):
            if i == 0:
                insert_row_excel_list(app, students_list_for_excel, course_name, '', time, moderator)
            else:
                insert_row_excel_list(app, students_list_for_excel, 'merge', 'merge', '', moderator)
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo inserisce la sezione relativa alla nota
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale    
def insert_note_section(app, students_list_for_excel):
    
    try:
        
        for i in range(0, 6):
            insert_row_excel_list(app, students_list_for_excel, TEXT_NOTE, TEXT_NOTE, TEXT_NO_BORDER, TEXT_NO_BORDER)
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per inserire le righe all'interno della lista per excel
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale   
# text_cell1 -> testo della cella 1
# text_cell2 -> testo della cella 2
# text_cell3 -> testo della cella 3
# text_cell4 -> testo della cella 4
def insert_row_excel_list(app, students_list_for_excel, text_cell1, text_cell2, text_cell3, text_cell4):
    
    try:
        students_list_for_excel['Nome e Cognome'].append(text_cell1)
        students_list_for_excel['Entrata'].append(text_cell2)
        students_list_for_excel['Uscita'].append(text_cell3)
        students_list_for_excel['Minuti'].append(text_cell4)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo serve per inserire la parte finale relative alla firma e ai totali
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale   
def insert_totals_and_sign(app, students_list_for_excel):
    
    try:
        
        insert_row_excel_list(app, students_list_for_excel, TEXT_TOTALE_PRESENZE_GIORNALIERE, TEXT_TOTALE_ASSENZE_GIORNALIERE, TEXT_TOTALE_ORE_GIORNALIERE, TEXT_FIRMA_TUTOR)
        insert_row_excel_list(app, students_list_for_excel, TEXT_N, TEXT_N, TEXT_N, TEXT_N)
    
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per la gestione della generazione del report
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale  
# index_without_border -> indici dove non deve essere applicato il bordo
# course_name -> nome del corso
def generate_excel_report(app, students_list_for_excel, course_name, time, moderators, filename):
    
    try:
        
        current_path = os.getcwd()
        path_generation_xlsx_report = current_path + '/generazione report personalizzato'
        
        if not os.path.exists(path_generation_xlsx_report):
            os.mkdir(path_generation_xlsx_report)
        
        excel_filename_and_path = f'{path_generation_xlsx_report}/{filename[(filename.rfind("/") + 1):]}'
        
        excel_df = pd.DataFrame(students_list_for_excel)
        
        excel_df.to_excel(excel_filename_and_path, index=False, sheet_name=course_name)
        
        wb = load_workbook(excel_filename_and_path)
        ws = wb.active
        
        ws.column_dimensions['A'].width = max([len(student) for student in excel_df['Nome e Cognome']]) + 5
        ws.column_dimensions['B'].width = max([len(date) for date in excel_df['Entrata']]) + 5
        ws.column_dimensions['C'].width = max([len(date) for date in excel_df['Uscita']]) + 10
        ws.column_dimensions['D'].width = max([len(str(minuti)) for minuti in excel_df['Minuti']]) + 10
        
        ws.row_dimensions[1].height = 30
        
        for column in COLUMN_REPORT_EXCEL:
            for cell in ws[column]:
                if cell.row == 1:
                    cell.font = Font(name='Verdana', bold=True)
                else:
                    cell.font = Font(name='Verdana')
        
        
        border = Border(
            left=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            right=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            top=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
            bottom=Side(border_style=BORDER_STYLE_THIN, color=BORDER_STYLE_COLOR_BLACK),
        )
        
        text_details_course_section_cells_merge = [TEXT_MATERIA, course_name, TEXT_MERGE, TEXT_MERGE, TEXT_MERGE]
        for text in text_details_course_section_cells_merge:
            merge_cells_excel(app, ws, 0, 2, text)
            
        merge_cells_excel(app, ws, 5, 2, TEXT_NOTE)
        set_bold_text_excel(app, ws, ['A'], TEXT_NO_BORDER)
        
        all_values_to_center = VALUES_TO_CENTER + [course_name, time] + list(moderators)
        
        set_value_position(app, ws, COLUMN_REPORT_EXCEL, all_values_to_center, HORIZONTAL_ALIGNMENT_CENTER)
        
        set_border_column_excel(app, ws, COLUMN_REPORT_EXCEL, border)
        
        wb.save(excel_filename_and_path)
        
        generate_modal(app, TEXT_REPORT_GENERATO_CON_SUCCESSO, 'Successo')
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo serve per inserire la dicitura 'no_border' all'interno delle celle al quale non si vuole assegnare il bordo
# app -> istanza CustomTKinter
# students_list_for_excel -> lista per lo sviluppo del report excel finale  
# number_row_without_border -> numero di righe senza bordo
def generate_row_without_border(app, students_list_for_excel, number_row_without_border):
    
    try:
        
        for i in range(0, number_row_without_border):
            insert_row_excel_list(app, students_list_for_excel, TEXT_NO_BORDER, TEXT_NO_BORDER, TEXT_NO_BORDER, TEXT_NO_BORDER)
            
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo serve per eseguire il merge delle righe e colonne in base ai valori inseriti nei parametri
# app -> istanza CustomTKinter       
# ws -> worksheet
# number_row -> numero di righe da mergiare (0 indica che si deve mergiare la stessa riga | es. troviamo valore alla riga 44: numero=riga: 0=44, 1=45, 2=46, ...)
# number_column -> numero di colonne da mergiare (es. numero=colonna : 1=A, 2=B, 3=C, ...)
# value_to_search -> valore da ricercare all'interno del worksheet
def merge_cells_excel(app, ws, number_row, number_column, value_to_search):
    
    try:
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value == value_to_search:
                    
                    if value_to_search == TEXT_MERGE:
                        cell.value = ''
                    
                    column_letter = cell.column_letter
                    row_number = cell.row
                    end_column_letter = get_column_letter(number_column)
                    end_row_merge = row_number + number_row
                    ws.merge_cells(f'{column_letter}{row_number}:{end_column_letter}{end_row_merge}')
                    break
                
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)
        
# il metodo serve per settare il valore bold alle colonne inserite fino al ritrovamento di un certo valore
# app -> istanza CustomTKinter       
# ws -> worksheet
# columns -> array contentente tutte le colonne al quale applicare il valore bold ['A', 'B', ...]
# values_end_bold -> il valore che indica quando bisogna smettere di applicare lo stile bold
def set_bold_text_excel(app, ws, columns, value_end_bold):
    
    try:
        
        for column in columns:
            for cell in ws[column]:
                
                if cell.value == value_end_bold:
                    break
                
                if cell.row > 1:
                    ws[f'{column}{cell.row}'].font = Font(name='Verdana', bold=True)
        
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo ritorna l'orario di inizio e fine corso
# app -> istanza CustomTKinter
# date_created_course -> data di creazione del corso
# date_ended_course -> data fine corso
def get_time_from_date_created_and_ended_course(app, date_created_course, date_ended_course):
    
    try:
        time_started_course = str(datetime.strptime(date_created_course, PATTERN_DD_MM_YY_WITH_TIME).time())
        time_ended_course = str(datetime.strptime(date_ended_course, PATTERN_DD_MM_YY_WITH_TIME).time())
        time = time_started_course[0:time_started_course.rfind(':')] + ' - ' + time_ended_course[0:time_ended_course.rfind(':')]
        return time
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo ritorna tutti i moderatori presenti
# app -> istanza CustomTKinter     
# moderators -> lista moderatori
def get_moderators_string(app, moderators):
    
    try:
        moderators_str = ''
        for moderator in moderators:
            moderators_str += moderator + ' - '
            
        moderators_str = moderators_str[0:moderators_str.rfind(' - ')]
        return moderators_str
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)

# il metodo si occupa di cambiare il posizionamento delle celle in base ai valori
# app -> istanza CustomTKinter
# ws -> worksheet
# all_values_to_change_position -> lista di valori al quale cambiare la posizione
# horizontal_position -> posizione orizzontale
# vertical_position -> posizione verticale
def set_value_position(app, ws, columns, all_values_to_change_position, horizontal_position = HORIZONTAL_ALIGNMENT_CENTER, vertical_position = None):
    
    try:
        
        for column in columns:
            for cell in ws[column]:
                if cell.value in all_values_to_change_position:
                    row = cell.row
                    if vertical_position != None:
                        ws[f'{column}{row}'].alignment = Alignment(horizontal=horizontal_position, vertical=vertical_position)
                    else:
                        ws[f'{column}{row}'].alignment = Alignment(horizontal=horizontal_position)
    except Exception as e:
        generate_modal(app, TEXT_MESSAGGIO_ERRORE)
        generate_log_errore(app, e)